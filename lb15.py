from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


def process_excel(input_file):
    wb = load_workbook(input_file)
    sheet = wb.active

    if "Результаты" in wb.sheetnames:
        result_sheet = wb["Результаты"]
    else:
        result_sheet = wb.create_sheet("Результаты")

    found = False

    for row in sheet.iter_rows(min_row=1, max_col=7, values_only=False):
        numbers = [cell.value for cell in row if isinstance(cell.value, int)]
        if len(numbers) != 7:
            continue

        # 2 числа повторяются дважды
        unique_counts = {num: numbers.count(num) for num in set(numbers)}
        double_repeats = [k for k, v in unique_counts.items() if v == 2]

        # Максимальное число не повторяется
        max_num = max(numbers)
        max_condition = numbers.count(max_num) == 1

        # Если оба условия выполнены
        if len(double_repeats) == 2 and max_condition:
            print(f"Строка: {numbers}")
            print(f"Сумма чисел: {sum(numbers)}")

            result_sheet.append(numbers)

            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for cell in row:
                cell.fill = fill
                cell.border = border

            found = True

        sum_cell = row[-1].offset(column=1)
        min_cell = row[-1].offset(column=2)

        sum_cell.value = sum(numbers)
        min_cell.value = min(numbers)

        sum_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        min_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        if found:
            break

    wb.save(input_file)
    print("Обработка завершена. Результаты сохранены в том же файле.")

input_excel = "24.xlsx"
process_excel(input_excel)
